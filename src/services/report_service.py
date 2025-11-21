# src/services/report_service.py
import os
import io
import datetime
import matplotlib.pyplot as plt
import pandas as pd
from collections import defaultdict
from typing import Dict, List, Any
import matplotlib

matplotlib.use('Agg')  # Para evitar problemas con GUI

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference

from src.config.settings import settings
from src.utils.logger import logger


class ReportService:
    def __init__(self):
        self.base_dir = settings.logging.base_dir
        self.reports_dir = os.path.join(self.base_dir, "reportes")
        self._ensure_directories()

    def _ensure_directories(self):
        """Asegurar que existan los directorios necesarios"""
        try:
            os.makedirs(self.reports_dir, exist_ok=True)

            # Crear estructura de año/mes/día
            today = datetime.datetime.now()
            year_dir = os.path.join(self.reports_dir, str(today.year))
            month_dir = os.path.join(year_dir, f"{today.month:02d}")
            day_dir = os.path.join(month_dir, f"{today.day:02d}")

            os.makedirs(day_dir, exist_ok=True)
            logger.info(f"📁 Directorios de reportes listos: {day_dir}")

        except Exception as e:
            logger.error(f"Error creando directorios: {str(e)}")

    def generate_usage_report(self, activity_records: Dict) -> Dict[str, Any]:
        """Generar reporte de uso del sistema"""
        try:
            if not activity_records:
                return {"summary": {"total_users": 0, "total_activities": 0, "avg_activities_per_user": 0}}

            # Procesar registros
            user_activities = defaultdict(int)
            action_breakdown = defaultdict(int)
            store_activities = defaultdict(int)
            hourly_usage = defaultdict(int)

            total_activities = 0
            unique_users = set()

            for user_id, records in activity_records.items():
                unique_users.add(user_id)
                user_activities[user_id] = len(records)
                total_activities += len(records)

                for record in records:
                    # Analizar tipo de acción
                    record_lower = record.lower()
                    if 'acción: store_access' in record_lower or 'tienda:' in record_lower:
                        action_breakdown['Acceso Tienda'] += 1
                        # Extraer código de tienda
                        if 'Tienda:' in record:
                            store_part = record.split('Tienda: ')[1]
                            store_code = store_part.split()[0] if ' ' in store_part else store_part
                            store_activities[store_code] += 1
                    elif 'acción: check_status' in record_lower or 'estado' in record_lower:
                        action_breakdown['Consulta Estado'] += 1
                    elif 'acción: audit' in record_lower or 'auditoria' in record_lower:
                        action_breakdown['Auditoría'] += 1
                    elif 'acción: reprint' in record_lower or 'reimpresion' in record_lower:
                        action_breakdown['Re-impresión'] += 1
                    elif 'acción: generate_image' in record_lower or 'imagen' in record_lower:
                        action_breakdown['Generar Imagen'] += 1
                    elif 'acción: comanda' in record_lower or 'comanda' in record_lower:
                        action_breakdown['Comanda'] += 1
                    elif 'acción: associated_code' in record_lower or 'código asociado' in record_lower:
                        action_breakdown['Código Asociado'] += 1
                    elif 'acción: start' in record_lower:
                        action_breakdown['Inicio Sesión'] += 1
                    else:
                        action_breakdown['Otras'] += 1

                    # Extraer hora de la actividad
                    try:
                        time_part = record.split(' - ')[0]
                        hour = int(time_part.split(' ')[1].split(':')[0])
                        hourly_usage[hour] += 1
                    except (IndexError, ValueError):
                        pass

            # Calcular métricas
            avg_activities = total_activities / len(unique_users) if unique_users else 0

            # Top tiendas
            top_stores = dict(sorted(store_activities.items(), key=lambda x: x[1], reverse=True)[:10])

            return {
                "summary": {
                    "total_users": len(unique_users),
                    "total_activities": total_activities,
                    "avg_activities_per_user": round(avg_activities, 2),
                    "analysis_period_days": 1,
                    "report_generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                },
                "action_breakdown": dict(action_breakdown),
                "top_stores": top_stores,
                "hourly_usage": dict(hourly_usage),
                "user_activities": dict(user_activities)
            }

        except Exception as e:
            logger.error(f"Error generando reporte de uso: {str(e)}")
            return {"summary": {"total_users": 0, "total_activities": 0, "avg_activities_per_user": 0}}

    def _create_chart_image(self, chart_type: str, data: Dict, title: str) -> io.BytesIO:
        """Crear imagen de gráfica para Excel"""
        try:
            plt.figure(figsize=(8, 6))

            if chart_type == 'pie':
                labels = list(data.keys())
                sizes = list(data.values())

                # Limitar etiquetas si son muchas
                if len(labels) > 8:
                    labels = labels[:8]
                    sizes = sizes[:8]
                    labels.append('Otros')
                    sizes.append(sum(list(data.values())[8:]))

                plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
                plt.axis('equal')

            elif chart_type == 'bar':
                labels = list(data.keys())
                values = list(data.values())

                plt.bar(labels, values, color='skyblue', edgecolor='navy')
                plt.xticks(rotation=45, ha='right')
                plt.ylabel('Cantidad')

            plt.title(title, fontsize=14, fontweight='bold')
            plt.tight_layout()

            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            buffer.seek(0)
            plt.close()

            return buffer

        except Exception as e:
            logger.error(f"Error creando gráfica {chart_type}: {str(e)}")
            # Gráfica de error
            plt.figure(figsize=(8, 6))
            plt.text(0.5, 0.5, f'Error en gráfica\n{str(e)}',
                     ha='center', va='center', transform=plt.gca().transAxes)
            plt.title(title)

            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            buffer.seek(0)
            plt.close()
            return buffer

    def generate_excel_report(self, activity_records: Dict, report_data: Dict, save_file: bool = True) -> io.BytesIO:
        """Generar reporte en Excel CON GRÁFICAS EMBEBIDAS"""
        try:
            buffer = io.BytesIO()
            workbook = Workbook()

            # Remover hoja por defecto
            workbook.remove(workbook.active)

            # ===== HOJA 1: RESUMEN EJECUTIVO =====
            ws_summary = workbook.create_sheet("Resumen Ejecutivo")
            ws_summary.sheet_view.showGridLines = False

            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            title_font = Font(size=16, bold=True, color="366092")

            # Título
            ws_summary.merge_cells('A1:F1')
            ws_summary['A1'] = "REPORTE DE USO - SISTEMA KFC BOT"
            ws_summary['A1'].font = title_font
            ws_summary['A1'].alignment = Alignment(horizontal='center')

            # Fecha de generación
            ws_summary['A3'] = "Fecha de generación:"
            ws_summary['B3'] = report_data['summary']['report_generated_at']
            ws_summary['A3'].font = Font(bold=True)

            # Métricas principales
            ws_summary['A5'] = "MÉTRICAS PRINCIPALES"
            ws_summary['A5'].font = Font(bold=True, size=12)

            metrics = [
                ("Total Usuarios", report_data['summary']['total_users']),
                ("Total Actividades", report_data['summary']['total_activities']),
                ("Promedio por Usuario", f"{report_data['summary']['avg_activities_per_user']:.2f}"),
                ("Período Analizado", f"{report_data['summary'].get('analysis_period_days', 1)} días")
            ]

            for i, (metric, value) in enumerate(metrics, start=6):
                ws_summary[f'A{i}'] = metric
                ws_summary[f'B{i}'] = value
                ws_summary[f'A{i}'].font = Font(bold=True)

            # ===== HOJA 2: DISTRIBUCIÓN DE ACTIVIDADES =====
            if report_data.get('action_breakdown'):
                ws_activities = workbook.create_sheet("Distribución Actividades")

                # Cabecera
                headers = ["Tipo de Actividad", "Cantidad", "Porcentaje"]
                for col, header in enumerate(headers, 1):
                    cell = ws_activities.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                # Datos
                total_actions = sum(report_data['action_breakdown'].values())
                sorted_activities = sorted(report_data['action_breakdown'].items(),
                                           key=lambda x: x[1], reverse=True)

                for row, (activity, count) in enumerate(sorted_activities, start=2):
                    percentage = (count / total_actions) * 100 if total_actions > 0 else 0

                    ws_activities.cell(row=row, column=1, value=activity)
                    ws_activities.cell(row=row, column=2, value=count)
                    ws_activities.cell(row=row, column=3, value=f"{percentage:.1f}%")

                # Gráfica de pastel
                if total_actions > 0:
                    chart_buffer = self._create_chart_image(
                        'pie', report_data['action_breakdown'],
                        'Distribución de Actividades'
                    )

                    # Guardar imagen temporal
                    temp_img_path = "temp_chart.png"
                    with open(temp_img_path, 'wb') as f:
                        f.write(chart_buffer.getvalue())

                    # Insertar imagen en Excel
                    img = XLImage(temp_img_path)
                    img.width = 400
                    img.height = 300
                    ws_activities.add_image(img, 'E2')

                    # Eliminar archivo temporal
                    os.remove(temp_img_path)

            # ===== HOJA 3: TOP TIENDAS =====
            if report_data.get('top_stores'):
                ws_stores = workbook.create_sheet("Top Tiendas")

                # Cabecera
                headers = ["Posición", "Tienda", "Actividades"]
                for col, header in enumerate(headers, 1):
                    cell = ws_stores.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                # Datos
                top_stores = list(report_data['top_stores'].items())[:15]  # Top 15
                for row, (store, count) in enumerate(top_stores, start=2):
                    ws_stores.cell(row=row, column=1, value=row - 1)
                    ws_stores.cell(row=row, column=2, value=store)
                    ws_stores.cell(row=row, column=3, value=count)

                # Gráfica de barras
                if top_stores:
                    stores_dict = dict(top_stores[:8])  # Máximo 8 para la gráfica
                    chart_buffer = self._create_chart_image(
                        'bar', stores_dict,
                        'Top Tiendas Más Activas'
                    )

                    # Guardar imagen temporal
                    temp_img_path = "temp_chart2.png"
                    with open(temp_img_path, 'wb') as f:
                        f.write(chart_buffer.getvalue())

                    # Insertar imagen en Excel
                    img = XLImage(temp_img_path)
                    img.width = 400
                    img.height = 300
                    ws_stores.add_image(img, 'E2')

                    os.remove(temp_img_path)

            # ===== HOJA 4: USO POR HORA =====
            if report_data.get('hourly_usage'):
                ws_hours = workbook.create_sheet("Uso por Hora")

                # Cabecera
                headers = ["Hora", "Actividades"]
                for col, header in enumerate(headers, 1):
                    cell = ws_hours.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                # Datos
                sorted_hours = sorted(report_data['hourly_usage'].items())
                for row, (hour, count) in enumerate(sorted_hours, start=2):
                    ws_hours.cell(row=row, column=1, value=f"{hour:02d}:00")
                    ws_hours.cell(row=row, column=2, value=count)

                # Gráfica de líneas
                if len(sorted_hours) > 1:
                    hours_dict = {f"{h:02d}:00": c for h, c in sorted_hours}
                    chart_buffer = self._create_chart_image(
                        'bar', hours_dict,
                        'Uso por Hora del Día'
                    )

                    temp_img_path = "temp_chart3.png"
                    with open(temp_img_path, 'wb') as f:
                        f.write(chart_buffer.getvalue())

                    img = XLImage(temp_img_path)
                    img.width = 400
                    img.height = 300
                    ws_hours.add_image(img, 'E2')

                    os.remove(temp_img_path)

            # ===== HOJA 5: REGISTROS DETALLADOS =====
            ws_details = workbook.create_sheet("Registros Detallados")

            # Cabecera
            headers = ["Usuario", "Fecha/Hora", "Actividad", "Tienda"]
            for col, header in enumerate(headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Datos (muestra de los últimos 500 registros)
            all_records = []
            for user_id, records in activity_records.items():
                for record in records:
                    # Parsear registro
                    parts = record.split(' - ')
                    if len(parts) >= 3:
                        fecha_hora = parts[0]
                        usuario_info = parts[1]
                        actividad = parts[2] if len(parts) > 2 else ""

                        # Extraer tienda si existe
                        tienda = ""
                        if 'Tienda:' in record:
                            tienda_part = record.split('Tienda: ')[1]
                            tienda = tienda_part.split(' - ')[0] if ' - ' in tienda_part else tienda_part

                        all_records.append({
                            'user_id': user_id,
                            'fecha_hora': fecha_hora,
                            'actividad': actividad,
                            'tienda': tienda
                        })

            # Ordenar por fecha (más recientes primero) y tomar muestra
            all_records.sort(key=lambda x: x['fecha_hora'], reverse=True)
            sample_records = all_records[:500]  # Máximo 500 registros

            for row, record in enumerate(sample_records, start=2):
                ws_details.cell(row=row, column=1, value=record['user_id'])
                ws_details.cell(row=row, column=2, value=record['fecha_hora'])
                ws_details.cell(row=row, column=3, value=record['actividad'])
                ws_details.cell(row=row, column=4, value=record['tienda'])

            # Ajustar anchos de columna
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar workbook
            workbook.save(buffer)
            buffer.seek(0)

            # Guardar archivo si se solicita
            if save_file:
                today = datetime.datetime.now()
                filename = f"reporte_completo_{today.strftime('%Y%m%d_%H%M')}.xlsx"
                filepath = os.path.join(self.reports_dir, str(today.year),
                                        f"{today.month:02d}", f"{today.day:02d}", filename)

                with open(filepath, 'wb') as f:
                    f.write(buffer.getvalue())

                logger.info(f"📊 Reporte Excel con gráficas guardado: {filepath}")
                buffer.seek(0)  # Reset buffer para Telegram

            return buffer

        except Exception as e:
            logger.error(f"Error generando Excel con gráficas: {str(e)}")
            # Devolver Excel básico como fallback
            return self._generate_basic_excel(activity_records, report_data)

    def _generate_basic_excel(self, activity_records: Dict, report_data: Dict) -> io.BytesIO:
        """Generar Excel básico como fallback"""
        try:
            buffer = io.BytesIO()

            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Hoja de resumen
                summary_data = {
                    'Métrica': ['Total Usuarios', 'Total Actividades', 'Promedio por Usuario', 'Fecha Generación'],
                    'Valor': [
                        report_data['summary']['total_users'],
                        report_data['summary']['total_activities'],
                        report_data['summary']['avg_activities_per_user'],
                        report_data['summary']['report_generated_at']
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Resumen', index=False)

                # Hoja de actividades
                if report_data.get('action_breakdown'):
                    action_df = pd.DataFrame(list(report_data['action_breakdown'].items()),
                                             columns=['Actividad', 'Cantidad'])
                    action_df.to_excel(writer, sheet_name='Actividades', index=False)

            buffer.seek(0)
            return buffer

        except Exception as e:
            logger.error(f"Error en Excel básico: {str(e)}")
            buffer = io.BytesIO()
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Error"
            ws['A1'] = f"Error generando reporte: {str(e)}"
            workbook.save(buffer)
            buffer.seek(0)
            return buffer

    # Mantener los otros métodos igual (generate_usage_chart, generate_detailed_txt_report, etc.)
    def generate_usage_chart(self, report_data: Dict, save_file: bool = True) -> io.BytesIO:
        """Generar gráficas de uso para Telegram"""
        try:
            if not report_data or not report_data.get('action_breakdown'):
                # Crear gráfica vacía
                fig, ax = plt.subplots(figsize=(10, 6))
                ax.text(0.5, 0.5, 'No hay datos suficientes\npara generar gráficas',
                        ha='center', va='center', transform=ax.transAxes, fontsize=14)
                ax.set_title('Gráficas de Uso - Sin Datos')

                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight', dpi=100)
                buffer.seek(0)
                plt.close()
                return buffer

            # Crear figura con subplots
            fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 10))
            fig.suptitle('Reporte de Uso - Sistema KFC Bot', fontsize=16, fontweight='bold')

            # Gráfica 1: Distribución de actividades
            action_data = report_data['action_breakdown']
            if action_data:
                actions = list(action_data.keys())
                counts = list(action_data.values())

                ax1.bar(actions, counts, color='skyblue', edgecolor='navy')
                ax1.set_title('Distribución por Tipo de Actividad')
                ax1.set_xlabel('Tipo de Actividad')
                ax1.set_ylabel('Cantidad')
                ax1.tick_params(axis='x', rotation=45)

            # Gráfica 2: Uso por hora
            hourly_data = report_data.get('hourly_usage', {})
            if hourly_data:
                hours = list(hourly_data.keys())
                counts = list(hourly_data.values())

                ax2.plot(hours, counts, marker='o', color='green', linewidth=2)
                ax2.set_title('Uso por Hora del Día')
                ax2.set_xlabel('Hora')
                ax2.set_ylabel('Actividades')
                ax2.grid(True, alpha=0.3)

            # Gráfica 3: Top tiendas
            store_data = report_data.get('top_stores', {})
            if store_data:
                stores = list(store_data.keys())[:8]
                counts = list(store_data.values())[:8]

                ax3.barh(stores, counts, color='lightcoral', edgecolor='darkred')
                ax3.set_title('Top Tiendas Más Activas')
                ax3.set_xlabel('Actividades')

            # Gráfica 4: Resumen general
            summary = report_data['summary']
            metrics = ['Usuarios', 'Actividades', 'Promedio']
            values = [summary['total_users'], summary['total_activities'], summary['avg_activities_per_user']]

            bars = ax4.bar(metrics, values, color=['lightgreen', 'lightblue', 'gold'])
            ax4.set_title('Métricas Generales')
            ax4.set_ylabel('Cantidad')

            # Agregar valores en las barras
            for bar, value in zip(bars, values):
                height = bar.get_height()
                ax4.text(bar.get_x() + bar.get_width() / 2., height,
                         f'{value}', ha='center', va='bottom')

            plt.tight_layout()

            # Guardar archivo si se solicita
            if save_file:
                today = datetime.datetime.now()
                filename = f"grafica_uso_{today.strftime('%Y%m%d_%H%M')}.png"
                filepath = os.path.join(self.reports_dir, str(today.year),
                                        f"{today.month:02d}", f"{today.day:02d}", filename)
                plt.savefig(filepath, bbox_inches='tight', dpi=150)
                logger.info(f"📊 Gráfica guardada: {filepath}")

            # Devolver buffer para Telegram
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', bbox_inches='tight', dpi=100)
            buffer.seek(0)
            plt.close()

            return buffer

        except Exception as e:
            logger.error(f"Error generando gráficas: {str(e)}")
            # Devolver gráfica de error
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.text(0.5, 0.5, f'Error generando gráficas:\n{str(e)}',
                    ha='center', va='center', transform=ax.transAxes, fontsize=12)
            ax.set_title('Error en Generación de Gráficas')

            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', bbox_inches='tight', dpi=100)
            buffer.seek(0)
            plt.close()
            return buffer

    def generate_detailed_txt_report(self, activity_records: Dict, report_data: Dict, save_file: bool = True) -> str:
        """Generar reporte detallado en texto"""
        try:
            summary = report_data['summary']

            report_text = [
                "=" * 60,
                "REPORTE DETALLADO - SISTEMA KFC BOT",
                "=" * 60,
                f"Fecha de generación: {summary['report_generated_at']}",
                f"Período analizado: {summary.get('analysis_period_days', 'N/A')} días",
                "",
                "RESUMEN EJECUTIVO:",
                f"  • Usuarios únicos: {summary['total_users']}",
                f"  • Total actividades: {summary['total_activities']}",
                f"  • Promedio por usuario: {summary['avg_activities_per_user']:.2f}",
                ""
            ]

            # Distribución por actividad
            if report_data.get('action_breakdown'):
                report_text.append("DISTRIBUCIÓN POR ACTIVIDAD:")
                total_actions = sum(report_data['action_breakdown'].values())
                for action, count in sorted(report_data['action_breakdown'].items(), key=lambda x: x[1], reverse=True):
                    percentage = (count / total_actions) * 100 if total_actions > 0 else 0
                    report_text.append(f"  • {action}: {count} ({percentage:.1f}%)")

            # Top tiendas
            if report_data.get('top_stores'):
                report_text.extend(["", "TOP TIENDAS MÁS ACTIVAS:"])
                for i, (store, count) in enumerate(list(report_data['top_stores'].items())[:10], 1):
                    report_text.append(f"  {i}. {store}: {count} actividades")

            report_text.extend([
                "",
                "=" * 60,
                "FIN DEL REPORTE",
                "=" * 60
            ])

            final_report = "\n".join(report_text)

            # Guardar archivo si se solicita
            if save_file:
                today = datetime.datetime.now()
                filename = f"reporte_detallado_{today.strftime('%Y%m%d_%H%M')}.txt"
                filepath = os.path.join(self.reports_dir, str(today.year),
                                        f"{today.month:02d}", f"{today.day:02d}", filename)

                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(final_report)

                logger.info(f"📊 Reporte TXT guardado: {filepath}")

            return final_report

        except Exception as e:
            error_msg = f"Error generando reporte detallado: {str(e)}"
            logger.error(error_msg)
            return error_msg

    def generate_daily_auto_report(self, activity_records: Dict) -> Dict:
        """Generar reporte automático diario"""
        try:
            logger.info("🤖 Generando reporte automático diario...")

            # Generar reporte completo
            report_data = self.generate_usage_report(activity_records)

            if report_data['summary']['total_activities'] == 0:
                logger.info("📊 No hay actividades para reporte automático")
                return report_data

            # Generar y guardar todos los formatos
            self.generate_usage_chart(report_data, save_file=True)
            self.generate_excel_report(activity_records, report_data, save_file=True)
            self.generate_detailed_txt_report(activity_records, report_data, save_file=True)

            logger.info("✅ Reporte automático diario completado")
            return report_data

        except Exception as e:
            logger.error(f"❌ Error en reporte automático: {str(e)}")
            return {"summary": {"error": str(e)}}